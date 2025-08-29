import os
import site
import sys 
import openpyxl
from datetime import date

import selenium
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC 


# to load the excel workbook with its path
file_path = r'C:/Users/Downloads/billings.xlsx'
excel_workbook = openpyxl.load_workbook(file_path)
# to identify active worksheet
excel_worksheet = excel_workbook.active

# do the check for all the RF codes in the A column 
print ("number of RF codes in total: ")
print (excel_worksheet.max_row - 1)    

# Initialize the WebDriver (in this case, Firefox)
driver = webdriver.Firefox()

# Open DEH HTML page
driver.get("https://mydei.dei.gr/el")

# iterate till the count of occupied rows
for m in range ( 2, excel_worksheet.max_row + 1):

    # the first time you visit the webpage, click on the rejection cookies button
    if (m==2):
        cookies_button = WebDriverWait(driver,10).until(
        EC.element_to_be_clickable((By.XPATH, "//button[@id='onetrust-reject-all-handler']"))
        )
        cookies_button.click()

    # click on the "Κοινόχρηστα"  (Shared Areas)
    koinoxrhsta_button = WebDriverWait(driver,3).until(
        EC.element_to_be_clickable((By.XPATH, "//section/div[5]/div[2]/div/div/div[2]/div/div/div[2]/button[3]/div[2]"))
    )
    driver.execute_script("arguments[0].click();", koinoxrhsta_button)

    # find Contract Account number field in the page
    contractAccount_input = driver.find_element(By.ID, "ContractAccount")
    # clear field
    contractAccount_input.clear() 
    # enter Contract Account number into the field
    contractAccount_input.send_keys(excel_worksheet.cell (row=m, column=1).value)    # column = 1 cause just the first column on excel has the RF codes

    # click on the "Είσοδος" (Login) button
    eisodos_button = WebDriverWait(driver,10).until(
        EC.element_to_be_clickable((By.XPATH, "//section/div[5]/div[2]/div/div/div[2]/div/div/div/form/div[2]/button"))
    )
    driver.execute_script("arguments[0].click();", eisodos_button)


    # click on the "ΕΚΔΟΣΗ ΛΟΓΑΡΙΑΣΜΩΝ" (Bills) span button 
    ekdoshLogariasmwn = WebDriverWait(driver,10).until(
        EC.element_to_be_clickable((By.XPATH, "//div[1]/div/div/div/section[3]/div/ul/li[2]/div/label/span"))
    )
    driver.execute_script("arguments[0].click();", ekdoshLogariasmwn)


    # find the top row of the bills table to check whether it's a new bill or not
    # top_table_row = driver.find_element(By.XPATH, "//table/tbody/tr[1]/td[3]")
    top_table_row = driver.find_element(By.XPATH, "//html/body/div[1]/div/div/div/section[3]/div/div[2]/div[1]/table/tbody/tr[1]/td[3]")


    # if the given RF code is new and there is no data saved for it, save the first date and make the "new bill?" column on the excel as F (we are considering that it has already been billed from the office)
    if (excel_worksheet.cell (row=m, column=2).value == None):
        # add the new date to the excel file
        excel_worksheet.cell (row=m, column=2).value = top_table_row.text
        excel_worksheet.cell (row=m, column=3).value = "F"  # if the third column is false it means that no new bill has been issued 
        # to save the workbook in location
        excel_workbook.save (file_path)
    
    # if the first date of the website is the same as the last saved one in the excel, then there is no new DEH bill
    elif (top_table_row.text==excel_worksheet.cell (row=m, column=2).value):    # the second column has the last billing period which the bill has been issued
        print("SAME DATE: NO BILLING UPDATE ON THE EXCEL FILE!")
        excel_worksheet.cell (row=m, column=3).value = "F"   # if the third column is false it means that no new bill has been issued 
        # to save the workbook in location
        excel_workbook.save (file_path)
        
    # if the first date of the website is different from the last saved one in the excel, then there is a new DEH bill => update the excel file   
    else:
        print ("NEW BILLING for the number: ")
        print (excel_worksheet.cell (row=m, column=1).value)
        # add the new date to the excel file
        excel_worksheet.cell (row=m, column=2).value = top_table_row.text
        excel_worksheet.cell (row=m, column=3).value = "T"  # if the third column is true it means that a new bill has been issued 
        # to save the workbook in location
        excel_workbook.save (file_path)


    # click on the "Επιστροφή" (Return) button
    epistrofh_button = driver.find_element(By.XPATH, "//div[1]/div/div/div/section[1]/a")
    epistrofh_button.click()

# Close the browser
driver.quit()